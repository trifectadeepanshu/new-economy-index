"""
Deprecated.

This script used an older listing-date composition model and an old constituent
universe. It is intentionally stopped before execution so it cannot generate a
workbook that disagrees with the app's quarterly top-50 index engine.

Use `npm run import:capiq -- --workbook "/path/to/workbook.xlsx"` to load the
canonical CapIQ workbook cache into the database and recompute the app index.
"""

raise SystemExit(
    "scripts/generate_excel.py is deprecated; use scripts/import-capiq-workbook.ts instead."
)

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2


def _load_database_url():
    """DATABASE_URL from the environment, falling back to .env.local."""
    url = os.environ.get("DATABASE_URL")
    if url:
        return url
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env.local")
    try:
        with open(env_path, "r", encoding="utf-8") as fh:
            for line in fh:
                if line.startswith("DATABASE_URL="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    except FileNotFoundError:
        pass
    raise SystemExit("DATABASE_URL not set (env or .env.local)")


DB_URL = _load_database_url()

INDEX_BASE_VALUE = 1000
CRORE = 10_000_000  # ₹1 crore = 1e7

PORTFOLIO_TICKERS = {
    "MEESHO", "URBANCO", "IXIGO", "BLUESTONE", "MOBIKWIK",
    "KISSHT", "BLACKBUCK", "IDEAFORGE", "SHADOWFAX", "NEPHROPLUS",
}

# Canonical constituents — keep in sync with lib/companies.ts (54 companies).
COMPANIES = [
    # (name, ticker, sector, listed_date)
    ("Eternal (Zomato)", "ETERNAL", "Platforms", "2021-07-23"),
    ("Swiggy", "SWIGGY", "Platforms", "2024-11-13"),
    ("PB Fintech (PolicyBazaar)", "POLICYBZR", "Platforms", "2021-11-15"),
    ("Meesho", "MEESHO", "Platforms", "2025-06-11"),
    ("Nykaa", "NYKAA", "Platforms", "2021-11-10"),
    ("Urban Company", "URBANCO", "Platforms", "2025-09-17"),
    ("Info Edge (Naukri)", "NAUKRI", "Platforms", "2006-11-21"),
    ("CarTrade.com", "CARTRADE", "Platforms", "2021-08-20"),
    ("ixigo", "IXIGO", "Platforms", "2024-06-18"),
    ("Yatra Online", "YATRA", "Platforms", "2023-09-21"),
    ("FirstCry", "FIRSTCRY", "Consumer Brands", "2024-08-13"),
    ("Indigo Paints", "INDIGOPNTS", "Consumer Brands", "2021-01-22"),
    ("Go Colors", "GOCOLORS", "Consumer Brands", "2021-11-30"),
    ("Sula Vineyards", "SULA", "Consumer Brands", "2022-12-22"),
    ("Wakefit", "WAKEFIT", "Consumer Brands", "2025-07-10"),
    ("Ola Electric", "OLAELEC", "Consumer Brands", "2024-08-09"),
    ("Bikaji Foods", "BIKAJI", "Consumer Brands", "2022-11-07"),
    ("Mamaearth (Honasa)", "HONASA", "Consumer Brands", "2023-11-07"),
    ("BlueStone Jewellery", "BLUESTONE", "Consumer Brands", "2025-06-25"),
    ("Ather Energy", "ATHERENERG", "Consumer Brands", "2025-05-06"),
    ("Lenskart", "LENSKART", "Consumer Brands", "2025-10-08"),
    ("Paytm", "PAYTM", "Fintech", "2021-11-18"),
    ("Groww", "GROWW", "Fintech", "2025-06-12"),
    ("Pine Labs", "PINELABS", "Fintech", "2025-07-24"),
    ("Zaggle", "ZAGGLE", "Fintech", "2023-09-22"),
    ("MobiKwik", "MOBIKWIK", "Fintech", "2024-12-18"),
    ("Kissht (OnEMI)", "KISSHT", "Fintech", "2026-05-08"),
    ("Go Digit Insurance", "GODIGIT", "Fintech", "2024-05-23"),
    ("Aadhar Housing Finance", "AADHARHFC", "Fintech", "2024-05-15"),
    ("Five Star Business Finance", "FIVESTAR", "Fintech", "2022-11-21"),
    ("HomeFirst Finance", "HOMEFIRST", "Fintech", "2022-02-03"),
    ("India Shelter Finance", "INDIASHLTR", "Fintech", "2023-12-20"),
    ("Northern Arc Capital", "NORTHARC", "Fintech", "2024-09-19"),
    ("AYE Finance", "AYE", "Fintech", "2026-02-26"),
    ("Delhivery", "DELHIVERY", "B2B", "2022-05-24"),
    ("IndiaMart", "INDIAMART", "B2B", "2019-07-04"),
    ("TBO.com", "TBOTEK", "B2B", "2024-05-15"),
    ("BlackBuck", "BLACKBUCK", "B2B", "2024-11-22"),
    ("Awfis", "AWFIS", "B2B", "2024-05-30"),
    ("IdeaForge", "IDEAFORGE", "B2B", "2023-06-26"),
    ("IndiQube", "INDIQUBE", "B2B", "2024-08-19"),
    ("UniCommerce", "UNIECOM", "B2B", "2024-08-13"),
    ("Shadowfax", "SHADOWFAX", "B2B", "2026-01-28"),
    ("MapMyIndia", "MAPMYINDIA", "SaaS", "2021-12-21"),
    ("Fractal Analytics", "FRACTAL", "SaaS", "2026-01-14"),
    ("Capillary Technologies", "CAPILLARY", "SaaS", "2025-02-18"),
    ("Amagi", "AMAGI", "SaaS", "2026-01-22"),
    ("RateGain", "RATEGAIN", "SaaS", "2021-12-17"),
    ("Tracxn Technologies", "TRACXN", "SaaS", "2022-10-20"),
    ("Medi Assist", "MEDIASSIST", "Healthcare", "2024-02-15"),
    ("NephroPlus (Nephrocare Health Services)", "NEPHROPLUS", "Healthcare", "2025-12-17"),
    ("Physics Wallah", "PWL", "Edtech", "2025-11-13"),
    ("Nazara Technologies", "NAZARA", "Gaming", "2021-03-30"),
    ("SEDEMAC Mechatronics", "SEDEMAC", "Deep Tech", "2026-03-11"),
]

NAME_BY_TICKER = {t: n for n, t, _, _ in COMPANIES}
SECTOR_BY_TICKER = {t: s for _, t, s, _ in COMPANIES}
MEMBERS = [(t, ld) for _, t, _, ld in COMPANIES]
PORTFOLIO_MEMBERS = [(t, ld) for _, t, _, ld in COMPANIES if t in PORTFOLIO_TICKERS]

# ── palette ──────────────────────────────────────────────────────────────────
NAVY, ORANGE, WHITE, LIGHT = "0F2040", "E07A38", "FFFFFF", "F2F4F8"
GREEN, RED, GREY = "1F8A4C", "C0392B", "8A9BB5"
PCT_FMT = '+0.00%;-0.00%;0.00%'


def hex_fill(c):
    return PatternFill("solid", fgColor=c)


def bold(size=11, color=WHITE):
    return Font(bold=True, size=size, color=color)


def thin_border():
    s = Side(style="thin", color="D1D9E6")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_row(ws, row, cols, labels):
    for c, label in zip(cols, labels):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = hex_fill(NAVY)
        cell.font = bold(10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── data ─────────────────────────────────────────────────────────────────────

def fetch_all_prices(cur):
    cur.execute("SELECT date::text, ticker, close_price::float FROM stock_snapshots ORDER BY date ASC")
    by_date = {}
    for dt, tk, px in cur.fetchall():
        by_date.setdefault(dt, {})[tk] = px
    return by_date


def fetch_shares(cur):
    cur.execute("""
        SELECT DISTINCT ON (ticker) ticker, shares::float
        FROM share_counts ORDER BY ticker, quarter_end DESC
    """)
    return {tk: sh for tk, sh in cur.fetchall()}


def fetch_latest_dates(cur):
    cur.execute("""
        SELECT DISTINCT ON (ticker) ticker, date::text
        FROM stock_snapshots ORDER BY ticker, date DESC
    """)
    return {tk: dt for tk, dt in cur.fetchall()}


# ── divisor engine (mirrors lib/index-engine.ts) ──────────────────────────────

def compute_index(prices_by_date, shares, members, base_value=INDEX_BASE_VALUE):
    """Return (rows, divisor, composition, last_close).

    rows: list of {date, value, num, chg_frac, mc, weights{ticker: w}}
    chg_frac is a day-over-day fraction (use a '%' number format to display).
    """
    dates = sorted(prices_by_date.keys())
    last_close = {}
    active, active_sig, divisor, seen, prev = [], "", 0.0, False, None
    rows = []

    def total_mc(comp):
        return sum(last_close[t] * shares[t] for t in comp if t in last_close and t in shares)

    for d in dates:
        for tk, px in prices_by_date[d].items():
            last_close[tk] = px

        eligible = [t for (t, ld) in members if ld <= d and t in last_close and t in shares]
        sig = ",".join(sorted(eligible))

        if eligible and sig != active_sig:
            new_mc = total_mc(eligible)
            if new_mc > 0:
                if not seen:
                    divisor, seen = new_mc / base_value, True
                else:
                    level_before = total_mc(active) / divisor if (active and divisor > 0) else base_value
                    if level_before > 0:
                        divisor = new_mc / level_before
                active, active_sig = eligible, sig

        if not seen or divisor <= 0:
            continue
        mc = total_mc(active)
        if mc <= 0:
            continue

        value = mc / divisor
        chg = (value / prev - 1) if prev else None
        weights = {t: (last_close[t] * shares[t]) / mc for t in active}
        rows.append({"date": d, "value": value, "num": len(active), "chg": chg, "mc": mc, "weights": weights})
        prev = value

    return rows, divisor, active, dict(last_close)


# ── Sheet 1: NEI Constituents ─────────────────────────────────────────────────

def build_constituents_sheet(ws, rows, divisor, composition, last_close, shares, latest_dates):
    ws.title = "NEI Constituents"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 32

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "Trifecta New Economy Index — Constituent Snapshot"
    t.fill = hex_fill(NAVY)
    t.font = Font(bold=True, size=13, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Company", "Ticker", "Sector", "Portfolio?", "Listed Date",
        "Latest Date", "Latest Price (₹)", "Shares (Cr)", "Market Cap (₹ Cr)", "Weight %",
    ]
    apply_header_row(ws, 2, range(1, 11), headers)

    last = rows[-1] if rows else {"mc": 0.0, "weights": {}, "value": INDEX_BASE_VALUE}
    total_mc = last["mc"]

    members = sorted(
        (tk for tk in composition if tk in last_close and tk in shares),
        key=lambda tk: last_close[tk] * shares[tk],
        reverse=True,
    )

    for i, tk in enumerate(members):
        r = i + 3
        price, sh = last_close[tk], shares[tk]
        mcap = price * sh
        weight = last["weights"].get(tk, 0.0)
        is_pf = tk in PORTFOLIO_TICKERS
        row_fill = hex_fill("EEF2F8") if i % 2 == 0 else hex_fill(WHITE)

        data = [
            NAME_BY_TICKER.get(tk, tk), tk, SECTOR_BY_TICKER.get(tk, ""),
            "✓" if is_pf else "", next(ld for t2, ld in MEMBERS if t2 == tk),
            latest_dates.get(tk, ""), price, sh / CRORE, mcap / CRORE, weight,
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="left" if c in (1, 3) else "center" if c in (2, 4, 5, 6) else "right",
                vertical="center",
            )
            if c == 4 and is_pf:
                cell.font = Font(bold=True, color=ORANGE)
            if c == 7:
                cell.number_format = '₹#,##0.00'
            if c in (8, 9):
                cell.number_format = '#,##0'
            if c == 10:
                cell.number_format = '0.00%'

    n = len(members)
    r = n + 3
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"].value = f"Total Market Cap (₹ Cr)  ·  Divisor = {divisor:,.0f}  ·  n = {n}"
    ws[f"A{r}"].fill = hex_fill(NAVY)
    ws[f"A{r}"].font = bold(10)
    ws[f"A{r}"].alignment = Alignment(horizontal="right", vertical="center")

    mc_cell = ws.cell(row=r, column=9, value=total_mc / CRORE)
    mc_cell.fill = hex_fill(NAVY)
    mc_cell.font = bold(10)
    mc_cell.number_format = '#,##0'
    mc_cell.alignment = Alignment(horizontal="right", vertical="center")

    iv_cell = ws.cell(row=r, column=10, value=last["value"])
    iv_cell.fill = hex_fill(ORANGE)
    iv_cell.font = bold(11)
    iv_cell.number_format = '#,##0.00'
    iv_cell.alignment = Alignment(horizontal="right", vertical="center")

    for c, w in enumerate([38, 12, 16, 11, 12, 12, 14, 12, 16, 11], 1):
        col_width(ws, c, w)
    for i in range(n + 1):
        ws.row_dimensions[i + 3].height = 18
    ws.freeze_panes = "C3"

    note_row = n + 5
    ws.merge_cells(f"A{note_row}:J{note_row}")
    ws[f"A{note_row}"].value = (
        "Methodology: market-cap weighted divisor index. Weight = company market cap ÷ total index "
        "market cap. Index Value = total market cap ÷ divisor (= 1,000 at the 1 March 2021 inception). "
        "Shares are constant point-in-time counts; market cap moves with price."
    )
    ws[f"A{note_row}"].font = Font(italic=True, size=9, color=GREY)
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center")


# ── Sheet 2: NEI Index History ────────────────────────────────────────────────

def build_history_sheet(ws, rows):
    ws.title = "NEI Index History"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 32

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "Trifecta New Economy Index — Daily History"
    t.fill = hex_fill(NAVY)
    t.font = Font(bold=True, size=13, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Date", "Index Value", "Daily Chg %", "No. of Companies", "Since Inception %"]
    apply_header_row(ws, 2, range(1, 6), headers)

    for i, row in enumerate(rows):
        r = i + 3
        row_fill = hex_fill("EEF2F8") if i % 2 == 0 else hex_fill(WHITE)
        since = row["value"] / INDEX_BASE_VALUE - 1

        for c, val in enumerate([row["date"], row["value"], row["chg"], row["num"], since], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if c == 1 else "right", vertical="center")

        ws.cell(row=r, column=2).number_format = '#,##0.00'
        chg_cell = ws.cell(row=r, column=3)
        chg_cell.number_format = PCT_FMT
        if row["chg"] is not None:
            chg_cell.font = Font(color=GREEN if row["chg"] >= 0 else RED, bold=True, size=10)
        si_cell = ws.cell(row=r, column=5)
        si_cell.number_format = PCT_FMT
        si_cell.font = Font(color=GREEN if since >= 0 else RED, bold=True, size=10)

    n = len(rows)
    last = rows[-1] if rows else {"date": None, "value": INDEX_BASE_VALUE}
    tr = n + 3
    ws.merge_cells(f"A{tr}:C{tr}")
    ws[f"A{tr}"].value = f"Latest  ({last['date']})"
    ws[f"A{tr}"].fill = hex_fill(NAVY)
    ws[f"A{tr}"].font = bold(10)
    ws[f"A{tr}"].alignment = Alignment(horizontal="right", vertical="center")

    iv = ws.cell(row=tr, column=4, value=last["value"])
    iv.fill = hex_fill(ORANGE)
    iv.font = bold(11)
    iv.number_format = '#,##0.00'
    iv.alignment = Alignment(horizontal="right", vertical="center")

    si_total = last["value"] / INDEX_BASE_VALUE - 1
    sc = ws.cell(row=tr, column=5, value=si_total)
    sc.fill = hex_fill(NAVY)
    sc.font = bold(10, color=GREEN if si_total >= 0 else RED)
    sc.number_format = PCT_FMT
    sc.alignment = Alignment(horizontal="right", vertical="center")

    for c, w in enumerate([14, 14, 14, 18, 18], 1):
        col_width(ws, c, w)
    for i in range(n + 1):
        ws.row_dimensions[i + 3].height = 16
    ws.freeze_panes = "B3"

    note_row = n + 5
    ws.merge_cells(f"A{note_row}:E{note_row}")
    ws[f"A{note_row}"].value = (
        "Note: market-cap weighted divisor index. Daily Chg % is the day-over-day change in the index "
        "level. Since Inception % = (Index Value ÷ 1,000 − 1). Base: 1,000 on 1 March 2021."
    )
    ws[f"A{note_row}"].font = Font(italic=True, size=9, color=GREY)
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center")


# ── Sheet 3: Portfolio Sub-Index ──────────────────────────────────────────────

def build_portfolio_sheet(ws, rows, composition):
    ws.title = "Portfolio Sub-Index"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14

    if not rows:
        ws["A1"].value = "No portfolio data found."
        return

    # Portfolio tickers that ever appear, ordered by latest weight (largest first).
    last_weights = rows[-1]["weights"]
    tickers = sorted(
        {t for row in rows for t in row["weights"]},
        key=lambda t: last_weights.get(t, 0.0),
        reverse=True,
    )
    n_tickers = len(tickers)
    last_col = 4 + n_tickers  # Date | Level | Chg % | # Active | weights…

    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    t = ws["A1"]
    t.value = "Trifecta Portfolio Sub-Index — Market-Cap Weighted Divisor Index"
    t.fill = hex_fill(NAVY)
    t.font = Font(bold=True, size=13, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 46

    fixed = ["Date", "Index Level", "Daily Chg %", "# Active"]
    ticker_headers = [f"{NAME_BY_TICKER.get(t, t)}\n({t}) wt %" for t in tickers]
    apply_header_row(ws, 3, range(1, last_col + 1), fixed + ticker_headers)
    for c in range(5, last_col + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = hex_fill("1A3A6B")
        cell.font = Font(bold=True, size=9, color=ORANGE)

    for i, row in enumerate(rows):
        r = i + 4
        row_fill = hex_fill("EEF2F8") if i % 2 == 0 else hex_fill(WHITE)

        c1 = ws.cell(row=r, column=1, value=row["date"])
        c1.fill = row_fill
        c1.border = thin_border()
        c1.alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws.cell(row=r, column=2, value=row["value"])
        c2.fill = hex_fill("E8F0FB") if i % 2 == 0 else hex_fill("F0F5FF")
        c2.border = thin_border()
        c2.number_format = '#,##0.00'
        c2.font = Font(bold=True, size=10, color=NAVY)
        c2.alignment = Alignment(horizontal="right", vertical="center")

        c3 = ws.cell(row=r, column=3, value=row["chg"])
        c3.fill = row_fill
        c3.border = thin_border()
        if row["chg"] is not None:
            c3.number_format = PCT_FMT
            c3.font = Font(color=GREEN if row["chg"] >= 0 else RED, size=10, bold=True)
        c3.alignment = Alignment(horizontal="right", vertical="center")

        c4 = ws.cell(row=r, column=4, value=row["num"])
        c4.fill = row_fill
        c4.border = thin_border()
        c4.alignment = Alignment(horizontal="center", vertical="center")

        for j, tk in enumerate(tickers):
            w = row["weights"].get(tk)
            cell = ws.cell(row=r, column=5 + j)
            cell.fill = row_fill
            cell.border = thin_border()
            if w is not None:
                cell.value = w
                cell.number_format = '0.0%'
                cell.font = Font(color=NAVY, size=9)
            else:
                cell.value = "—"
                cell.font = Font(color=GREY, size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")

    for c, w in [(1, 13), (2, 14), (3, 14), (4, 10)]:
        col_width(ws, c, w)
    for j in range(n_tickers):
        col_width(ws, 5 + j, 13)
    for i in range(len(rows)):
        ws.row_dimensions[i + 4].height = 16
    ws.freeze_panes = "B4"

    note_row = len(rows) + 6
    ws.merge_cells(f"A{note_row}:{get_column_letter(last_col)}{note_row}")
    ws[f"A{note_row}"].value = (
        "Methodology: market-cap weighted divisor sub-index over the 10 portfolio companies, based at "
        "1,000 on the first portfolio IPO (IdeaForge, Jun 2023). Each company joins on its listing day "
        "via a divisor adjustment, keeping the level continuous. Weight % = company market cap ÷ "
        "portfolio market cap that day. The NEI keeps its own 1 March 2021 inception base."
    )
    ws[f"A{note_row}"].font = Font(italic=True, size=9, color=GREY)
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to database…")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    print("Fetching data…")
    prices_by_date = fetch_all_prices(cur)
    shares = fetch_shares(cur)
    latest_dates = fetch_latest_dates(cur)
    cur.close()
    conn.close()

    if not shares:
        raise SystemExit("share_counts is empty — run scripts/backfill-shares.ts first.")
    print(f"  {len(prices_by_date)} trading days · {len(shares)} tickers with shares")

    print("Computing divisor index…")
    nei_rows, divisor, composition, last_close = compute_index(prices_by_date, shares, MEMBERS)
    pf_rows, _, _, _ = compute_index(prices_by_date, shares, PORTFOLIO_MEMBERS)
    print(f"  NEI latest = {nei_rows[-1]['value']:,.2f} (n={nei_rows[-1]['num']}) · "
          f"Portfolio latest = {pf_rows[-1]['value']:,.2f} (n={pf_rows[-1]['num']})")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building Sheet 1: NEI Constituents…")
    build_constituents_sheet(wb.create_sheet(), nei_rows, divisor, composition, last_close, shares, latest_dates)
    print("Building Sheet 2: NEI Index History…")
    build_history_sheet(wb.create_sheet(), nei_rows)
    print("Building Sheet 3: Portfolio Sub-Index…")
    build_portfolio_sheet(wb.create_sheet(), pf_rows, composition)

    out = os.path.join(os.path.expanduser("~"), "Downloads", "NEI_Index_Model.xlsx")
    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    main()
